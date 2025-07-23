from flask import Flask, request, render_template_string, redirect, url_for, flash
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
import os
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

app = Flask(__name__)
app.secret_key = 'clave_secreta_segura'
EXCEL_FILE = 'datos.xlsx'

# ==== CONFIGURACIÓN CORREO ====
EMAIL_ORIGEN = 'tucorreo@gmail.com'      # Cambia por tu correo
EMAIL_CLAVE = 'clave_app_segura'         # Contraseña de aplicación si usas Gmail
EMAIL_DESTINO = 'jefeti@mariasalome.com' # Destinatario

SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 587

formulario_html = """
<!doctype html>
<html>
<head><title>Ventas</title></head>
<body>
<h2>Registro de Venta</h2>

{% with messages = get_flashed_messages() %}
  {% if messages %}
    <ul>
    {% for msg in messages %}
      <li><strong>{{ msg }}</strong></li>
    {% endfor %}
    </ul>
  {% endif %}
{% endwith %}

<form method="POST">
  <label>Cédula del Cliente:</label><br>
  <input type="text" name="cedula"><br><br>

  <label>Productos (formato: codigo:cantidad, separados por coma. Ej: P001:2,P002:3):</label><br>
  <textarea name="productos" rows="4" cols="50"></textarea><br><br>

  <input type="submit" value="Guardar">
</form>
</body>
</html>
"""

def cargar_datos():
    df_clientes = pd.read_excel(EXCEL_FILE, sheet_name='clientes')
    df_productos = pd.read_excel(EXCEL_FILE, sheet_name='productos')
    return df_clientes, df_productos

def enviar_correo(cedula, resumen):
    msg = MIMEMultipart()
    msg['From'] = EMAIL_ORIGEN
    msg['To'] = EMAIL_DESTINO
    msg['Subject'] = f"Venta registrada - Cliente {cedula}"

    cuerpo = f"""Se ha registrado una venta para el cliente con cédula {cedula}:

{resumen}

Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
    """

    msg.attach(MIMEText(cuerpo, 'plain'))

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_ORIGEN, EMAIL_CLAVE)
            server.send_message(msg)
        return True
    except Exception as e:
        print(f"Error enviando correo: {e}")
        return False

def guardar_venta_excel(cedula, productos):
    wb = load_workbook(EXCEL_FILE)
    if 'ventas' not in wb.sheetnames:
        ws = wb.create_sheet('ventas')
        ws.append(['cedula', 'referencia', 'cantidad', 'fecha'])
    else:
        ws = wb['ventas']

    df_productos = pd.read_excel(EXCEL_FILE, sheet_name='productos')

    resumen = ""
    # Validar stock
    for ref, cant in productos.items():
        disponible = df_productos.loc[df_productos['referencia'] == ref, 'cantidad_disponible'].values[0]
        if disponible < cant:
            return False, f"❌ No hay suficiente inventario de {ref} (hay {disponible}, pidió {cant})"

    # Registrar venta y actualizar stock
    for ref, cant in productos.items():
        idx = df_productos[df_productos['referencia'] == ref].index[0]
        df_productos.at[idx, 'cantidad_disponible'] -= cant
        ws.append([cedula, ref, cant, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        resumen += f"- Producto {ref}: {cant} unidades\n"

    # Guardar cambios
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_productos.to_excel(writer, sheet_name='productos', index=False)

    wb.save(EXCEL_FILE)

    # Enviar correo
    enviar_correo(cedula, resumen)

    return True, "✅ Venta registrada y correo enviado."

@app.route('/', methods=['GET', 'POST'])
def registrar_venta():
    if request.method == 'POST':
        cedula = request.form.get('cedula', '').strip()
        productos_raw = request.form.get('productos', '').strip()

        df_clientes, df_productos = cargar_datos()

        if cedula not in df_clientes['cedula'].astype(str).values:
            flash("⚠️ Cliente no encontrado.")
            return redirect(url_for('registrar_venta'))

        productos = {}
        try:
            for item in productos_raw.split(','):
                ref, cantidad = item.strip().split(':')
                ref = ref.strip()
                cantidad = int(cantidad.strip())
                if ref not in df_productos['referencia'].values:
                    flash(f"⚠️ Código de producto no válido: {ref}")
                    return redirect(url_for('registrar_venta'))
                productos[ref] = cantidad
        except Exception:
            flash("⚠️ Formato de productos incorrecto. Usa P001:2,P002:3")
            return redirect(url_for('registrar_venta'))

        ok, mensaje = guardar_venta_excel(cedula, productos)
        flash(mensaje)
        return redirect(url_for('registrar_venta'))

    return render_template_string(formulario_html)


	if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))  # Toma el puerto de la variable de entorno
    app.run(host='0.0.0.0', port=port, debug=True)
	
   if not os.path.exists(EXCEL_FILE):
        print(f"⛔ El archivo '{EXCEL_FILE}' no existe. Crea uno con hojas 'clientes' y 'productos'.")
    else:
        app.run(debug=True)